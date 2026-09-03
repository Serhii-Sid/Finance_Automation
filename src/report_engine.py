import os
import logging
from typing import Optional
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import (
    OUTPUT_FOLDER, OUTPUT_FILE, COL_DATE, COL_AMOUNT, COL_BALANCE, COL_ID, COL_CAT, COL_CARD, COL_DESC,
    COL_CLEARING_STATUS, COL_MCC
)
import datetime
from src.finance_logic import (
    expand_commission_splits_for_reports,
    detect_internal_transfers,
    process_cash_clearing,
    process_transit_vika,
    process_mono_investments,
    process_investment_transit_clearing,
    ReconciliationRegistry
)

logger = logging.getLogger(__name__)

def generate_daily_dashboard(df_ledger: pd.DataFrame) -> pd.DataFrame:
    """Генерує щоденний дашборд за структурою ручного леджера (Витрати.csv) з 4 візуальними секціями."""
    if df_ledger is None or df_ledger.empty:
        return pd.DataFrame(columns=[
            'Місяць', 'Дата', 'План', 'Витрати', 'Дохід', 
            'Різниця за день', 'Різниця за місяць', 'Різниця загалом'
        ])
    
    df = expand_commission_splits_for_reports(df_ledger)
    df[COL_DATE] = pd.to_datetime(df[COL_DATE])
    # Localize to naive just in case
    if df[COL_DATE].dt.tz is not None:
        df[COL_DATE] = df[COL_DATE].dt.tz_localize(None)
        
    df['Дата_Норм'] = df[COL_DATE].dt.normalize()
    min_date = df['Дата_Норм'].min()
    max_date = df['Дата_Норм'].max()
    
    # Generate continuous dates
    all_dates = pd.date_range(start=min_date, end=max_date, freq='D')
    
    # Витрати: сума < 0 (зберігаємо як від'ємні числа, ігноруємо перекази на власний рахунок, зняття готівки, транзит Віка та компенсовані інвестиції)
    expenses_mask = (df[COL_AMOUNT] < 0) & (~df[COL_CAT].isin(['переказ на власний рахунок', 'зняття готівки', 'транзит Віка', 'інвестиції (транзит Віка)']))
    df_expenses = df[expenses_mask].groupby('Дата_Норм')[COL_AMOUNT].sum()
    
    # Доходи: позитивні суми (Сума > 0, ігноруємо перекази з власного рахунку, поповнення готівкою та транзит Віка)
    income_mask = (df[COL_AMOUNT] > 0) & (~df[COL_CAT].isin(['переказ з власного рахунку', 'поповнення готівкою', 'транзит Віка']))
    df_income = df[income_mask].groupby('Дата_Норм')[COL_AMOUNT].sum()
    
    # Для закритих місяців рахуємо сумарні доходи по місяцях
    month_income_totals = {}
    for (yr, mn), group in df[income_mask].groupby([df[COL_DATE].dt.year, df[COL_DATE].dt.month]):
        month_income_totals[(yr, mn)] = group[COL_AMOUNT].sum()
        
    now = pd.Timestamp.now()
    plan_cache = {}
    
    ukr_months_nominative = {
        1: 'Січень', 2: 'Лютий', 3: 'Березень', 4: 'Квітень', 5: 'Травень', 6: 'Червень',
        7: 'Липень', 8: 'Серпень', 9: 'Вересень', 10: 'Жовтень', 11: 'Листопад', 12: 'Грудень'
    }
    
    rows = []
    for dt in all_dates:
        yr, mn = dt.year, dt.month
        days_in_month = pd.Period(year=yr, month=mn, freq='M').days_in_month
        
        # Дохідна логіка для дня
        daily_income = df_income.get(dt, 0.0)
        
        # План для дня
        if (yr, mn) not in plan_cache:
            is_closed = (yr < now.year) or (yr == now.year and mn < now.month)
            if is_closed:
                # План = Сума доходу за цей місяць / дні місяця
                tot_inc = month_income_totals.get((yr, mn), 0.0)
                plan_cache[(yr, mn)] = tot_inc / days_in_month
            else:
                # Відкритий місяць: прогноз 40000 / дні місяця
                plan_cache[(yr, mn)] = 40000.0 / days_in_month
                
        plan = plan_cache[(yr, mn)]
        daily_expense = df_expenses.get(dt, 0.0) # Вже від'ємне
        
        rows.append({
            'Місяць': f"{ukr_months_nominative[mn]} {yr}",
            'Дата': dt,
            'План': plan,
            'Витрати': daily_expense,
            'Дохід': daily_income
        })
        
    df_dash = pd.DataFrame(rows)
    
    # Різниця за день
    df_dash['Різниця за день'] = df_dash['План'] + df_dash['Витрати']
    
    # Різниця за місяць = cumulative sum of Різниця за день (reset every month)
    df_dash['Різниця за місяць'] = df_dash.groupby([df_dash['Дата'].dt.year, df_dash['Дата'].dt.month])['Різниця за день'].cumsum()
    
    # Різниця загалом = -48807.41 + cumulative sum of Різниця за день за весь період
    INITIAL_TOTAL_BALANCE = -48807.41
    df_dash['Різниця загалом'] = INITIAL_TOTAL_BALANCE + df_dash['Різниця за день'].cumsum()
    
    # Тимчасові стовпці для групування
    df_dash['temp_year'] = df_dash['Дата'].dt.year
    df_dash['temp_month'] = df_dash['Дата'].dt.month
    
    # Форматуємо 'Дата' як чистий рядок dd.mm.yyyy
    df_dash['Дата'] = df_dash['Дата'].dt.strftime('%d.%m.%Y')
    
    grouped = df_dash.groupby(['temp_year', 'temp_month'], sort=False)
    
    final_rows = []
    for (yr, mn), group in grouped:
        for idx, row in group.iterrows():
            final_rows.append(row.to_dict())
            
        # Розраховуємо суми за місяць
        sum_plan = group['План'].sum()
        sum_expenses = group['Витрати'].sum()
        sum_income = group['Дохід'].sum()
        
        # Беремо останні накопичувальні значення місяця
        last_day_row = group.iloc[-1]
        final_diff_month = last_day_row['Різниця за місяць']
        final_diff_total = last_day_row['Різниця загалом']
        
        # Створюємо підсумковий рядок РАЗОМ
        month_name = ukr_months_nominative[mn]
        rahom_row = {
            'Місяць': f"{month_name} {yr}",
            'Дата': f'РАЗОМ за {month_name} {yr}',
            'План': sum_plan,
            'Витрати': sum_expenses,
            'Дохід': sum_income,
            'Різниця за день': None,
            'Різниця за місяць': final_diff_month,
            'Різниця загалом': final_diff_total,
            'temp_year': yr,
            'temp_month': mn
        }
        final_rows.append(rahom_row)
        
    df_dash_final = pd.DataFrame(final_rows)
    df_dash_final = df_dash_final.drop(columns=['temp_year', 'temp_month'], errors='ignore')
    
    # Переконуємось у правильному порядку колонок
    col_order = ['Місяць', 'Дата', 'План', 'Витрати', 'Дохід', 'Різниця за день', 'Різниця за місяць', 'Різниця загалом']
    df_dash_final = df_dash_final[col_order]
    
    return df_dash_final

def rotate_outputs():
    """Керування версіями файлу Ledger."""
    base_name = "Total_Ledger_v{}.xlsx"
    for i in [2, 1]:
        old_v = os.path.join(OUTPUT_FOLDER, base_name.format(i))
        new_v = os.path.join(OUTPUT_FOLDER, base_name.format(i+1))
        if os.path.exists(old_v): os.replace(old_v, new_v)
    if os.path.exists(OUTPUT_FILE): 
        os.replace(OUTPUT_FILE, os.path.join(OUTPUT_FOLDER, base_name.format(1)))

def save_final_ledger(df: pd.DataFrame, df_dash: Optional[pd.DataFrame] = None, script_path: str = "", df_analytical: Optional[pd.DataFrame] = None):
    """Зберігає фінальний Excel з 5 листами (Total_Ledger, Income, Reconciliation_Audit, Expenses, Daily_Dashboard)."""
    try:
        rotate_outputs()

        from src.finance_logic import detect_internal_transfers, process_cash_clearing, ReconciliationRegistry

        # 1. Сирий недоторканий DataFrame для листа 'Total_Ledger'
        df_export = df.copy()
        df_export = df_export.drop(columns=[COL_CLEARING_STATUS], errors='ignore')
        df_export.insert(0, '№ п/п', range(1, len(df_export) + 1))
        
        ukr_months = {
            1: 'Січень', 2: 'Лютий', 3: 'Березень', 4: 'Квітень', 5: 'Травень', 6: 'Червень',
            7: 'Липень', 8: 'Серпень', 9: 'Вересень', 10: 'Жовтень', 11: 'Листопад', 12: 'Грудень'
        }
        df_export.insert(2, 'Місяць', df_export[COL_DATE].dt.month.map(ukr_months) + " " + df_export[COL_DATE].dt.year.astype(str))
        
        # 2. Формування df_analytical та збір аудиторського реєстру ReconciliationRegistry
        if df_analytical is None:
            df_analytical = detect_internal_transfers(df.copy())
            df_analytical = process_transit_vika(df_analytical)
            df_analytical = process_mono_investments(df_analytical)
            df_analytical = process_investment_transit_clearing(df_analytical)
            df_analytical = process_cash_clearing(df_analytical)

        # 3. Дашборд
        if df_dash is None or df_dash.empty:
            df_dash = generate_daily_dashboard(df_analytical)

        # 4. Підготовка експорту для вкладок Income та Expenses
        df_analytical_processed = expand_commission_splits_for_reports(df_analytical)
        df_analytical_export = df_analytical_processed.copy()
        df_analytical_export.insert(0, '№ п/п', range(1, len(df_analytical_export) + 1))
        df_analytical_export.insert(2, 'Місяць', df_analytical_export[COL_DATE].dt.month.map(ukr_months) + " " + df_analytical_export[COL_DATE].dt.year.astype(str))

        df_income = df_analytical_export[df_analytical_export[COL_AMOUNT] > 0].copy()
        df_expenses = df_analytical_export[df_analytical_export[COL_AMOUNT] < 0].copy()
        
        cols_to_drop_analysis = ['№ п/п', 'Місяць']
        df_income = df_income.drop(columns=cols_to_drop_analysis, errors='ignore')
        df_expenses = df_expenses.drop(columns=cols_to_drop_analysis, errors='ignore')

        desired_order = [COL_ID, COL_DATE, COL_CLEARING_STATUS, COL_CAT, COL_CARD, COL_DESC, COL_AMOUNT, COL_BALANCE, COL_MCC]
        income_cols = [c for c in desired_order if c in df_income.columns] + [c for c in df_income.columns if c not in desired_order]
        expenses_cols = [c for c in desired_order if c in df_expenses.columns] + [c for c in df_expenses.columns if c not in desired_order]
        df_income = df_income[income_cols]
        df_expenses = df_expenses[expenses_cols]

        # Заготовка порожньої структури для Reconciliation_Audit
        df_audit_header = pd.DataFrame(columns=[
            'Дата', 'Операція / ID', 'Картка / Опис',
            'Оригінальна сума', 'Скомпенсовано', 'Залишок (Витрати)', 'Деталі звірки'
        ])

        # Запис п'яти листів у суворо визначеному порядку:
        # 1. Total_Ledger, 2. Income, 3. Expenses, 4. Reconciliation_Audit, 5. Daily_Dashboard
        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', datetime_format='dd.mm.yyyy hh:mm:ss') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Total_Ledger')
            df_income.to_excel(writer, index=False, sheet_name='Income')
            df_expenses.to_excel(writer, index=False, sheet_name='Expenses')
            df_audit_header.to_excel(writer, index=False, sheet_name='Reconciliation_Audit')
            df_dash.to_excel(writer, index=False, sheet_name='Daily_Dashboard')

            # Стилізація листів
            for sheet_name in ['Total_Ledger', 'Income', 'Expenses', 'Reconciliation_Audit', 'Daily_Dashboard']:
                sheet = writer.sheets[sheet_name]
                sheet.views.sheetView[0].showGridLines = True
                
                if sheet_name in ('Total_Ledger', 'Income', 'Expenses'):
                    # Стандартна стилізація для основних таблиць
                    sheet.row_dimensions[1].height = 30
                    header_fill = PatternFill(start_color="5A5A5A", end_color="5A5A5A", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF", size=12)
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    thin_side = Side(style='thin', color="000000")
                    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                    for cell in sheet[1]:
                        if cell.value is not None:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = header_border

                    # Сітка та Зебра
                    thin_grid_side = Side(style='thin', color="A6A6A6")
                    thick_border_side = Side(style='medium', color="000000")
                    zebra_fill = PatternFill(start_color="E9E9E9", end_color="E9E9E9", fill_type="solid")
                    data_alignment = Alignment(vertical="center")

                    prev_period = None
                    group_start = 2
                    for row_idx in range(2, sheet.max_row + 1):
                        curr_period = None
                        is_new_period = False
                        if sheet_name == 'Total_Ledger':
                            curr_period = sheet.cell(row=row_idx, column=3).value
                            is_new_period = prev_period is not None and curr_period != prev_period
                        is_even = row_idx % 2 == 0

                        if is_new_period and sheet_name == 'Total_Ledger':
                            if row_idx - 1 > group_start:
                                sheet.row_dimensions.group(group_start + 1, row_idx - 1, outline_level=1)
                            group_start = row_idx

                        for col_idx in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            top_s = thick_border_side if (sheet_name == 'Total_Ledger' and is_new_period) else thin_grid_side
                            cell.border = Border(left=thin_grid_side, right=thin_grid_side, top=top_s, bottom=thin_grid_side)
                            
                            header_val = sheet.cell(row=1, column=col_idx).value
                            if header_val == 'Дата':
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            else:
                                cell.alignment = data_alignment
                            
                            if is_even:
                                cell.fill = zebra_fill
                        prev_period = curr_period

                    if sheet_name == 'Total_Ledger' and sheet.max_row > group_start:
                        sheet.row_dimensions.group(group_start + 1, sheet.max_row, outline_level=1)
                        sheet.sheet_properties.outlinePr.summaryBelow = False

                    sheet.freeze_panes = "A2"
                    sheet.auto_filter.ref = sheet.dimensions

                elif sheet_name == 'Reconciliation_Audit':
                    # --- Побудова деревоподібного листа звірки (Parent-Child Blocks) ---
                    sheet.row_dimensions[1].height = 30
                    header_fill_audit = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    header_font_audit = Font(bold=True, color="FFFFFF", size=11)
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    thin_grid_side = Side(style='thin', color="A6A6A6")
                    header_border = Border(left=thin_grid_side, right=thin_grid_side, top=thin_grid_side, bottom=thin_grid_side)

                    for cell in sheet[1]:
                        cell.font = header_font_audit
                        cell.fill = header_fill_audit
                        cell.alignment = header_alignment
                        cell.border = header_border

                    # Стилі блоків
                    parent_fill = PatternFill(fill_type=None)
                    parent_font = Font(bold=True, size=11, color="000000")

                    child_fill = PatternFill(fill_type=None)
                    child_font = Font(size=11, color="000000")

                    summary_fill = PatternFill(fill_type=None)
                    summary_font = Font(bold=True, size=11, color="000000")

                    twins_fill = PatternFill(fill_type=None)
                    twins_font = Font(size=11, color="000000")

                    align_center = Alignment(horizontal="center", vertical="center")
                    align_left = Alignment(horizontal="left", vertical="center")
                    align_right = Alignment(horizontal="right", vertical="center")

                    df_audit_data = ReconciliationRegistry.get_df()

                    events = []
                    if not df_audit_data.empty and 'Тип компенсації' in df_audit_data.columns:
                        df_cash = df_audit_data[df_audit_data['Тип компенсації'].str.contains('Cash Clearing|Готівка', case=False, na=False)].copy()
                        df_twins = df_audit_data[df_audit_data['Тип компенсації'].str.contains('Twins', case=False, na=False)].copy()
                        df_transit = df_audit_data[df_audit_data['Тип компенсації'].str.contains('Investment Transit|Віка', case=False, na=False)].copy()

                        # 1. Події Cash Clearing (групування за ID зняття)
                        if not df_cash.empty:
                            grouped_cash = df_cash.groupby('ID зняття', sort=False)
                            for w_id, group in grouped_cash:
                                first_w_date = pd.to_datetime(group.iloc[0]['Дата зняття'])
                                events.append({
                                    'type': 'cash',
                                    'date': first_w_date,
                                    'data': group
                                })

                        # 2. Події Twins (кожна транзакція — окрема подія)
                        if not df_twins.empty:
                            for _, t_row in df_twins.iterrows():
                                t_date = pd.to_datetime(t_row['Дата зняття'])
                                events.append({
                                    'type': 'twins',
                                    'date': t_date,
                                    'data': t_row
                                })

                        # 3. Події Investment Transit (групування за ID поповнення)
                        if not df_transit.empty:
                            grouped_transit = df_transit.groupby('ID поповнення', sort=False)
                            for dep_id, group in grouped_transit:
                                first_dep_date = pd.to_datetime(group.iloc[0]['Дата поповнення'])
                                events.append({
                                    'type': 'transit_vika',
                                    'date': first_dep_date,
                                    'data': group
                                })

                    # 3. Сортування всіх подій у порядку СПАДАННЯ (від найновіших до найстаріших)
                    events.sort(key=lambda x: x['date'], reverse=True)

                    current_row = 2

                    for event in events:
                        if event['type'] == 'cash':
                            group = event['data']
                            first_row = group.iloc[0]
                            w_date = first_row['Дата зняття']
                            w_id = str(first_row['ID зняття'])
                            w_desc = str(first_row.get('Опис зняття', '') or '')
                            w_card = str(first_row.get('Картка джерела', '') or w_desc).strip()
                            w_amount = float(first_row.get('Сума зняття', 0.0) or 0.0)

                            w_dt = pd.to_datetime(w_date) if not isinstance(w_date, (pd.Timestamp, datetime.datetime)) else w_date
                            w_date_str = w_dt.strftime('%d.%m') if pd.notnull(w_dt) else ""

                            # --- PARENT ROW ---
                            parent_desc = f"Зняття готівки в банкоматі ({w_card})" if w_card else "Зняття готівки в банкоматі"
                            parent_vals = [
                                w_date,
                                w_id,
                                parent_desc,
                                w_amount,
                                None,
                                None,
                                f"Оригінальне зняття готівки від {w_date_str}"
                            ]
                            sheet.row_dimensions[current_row].height = 24
                            for c_idx, val in enumerate(parent_vals, 1):
                                cell = sheet.cell(row=current_row, column=c_idx, value=val)
                                cell.font = parent_font
                                cell.fill = parent_fill
                                cell.border = header_border
                                if c_idx in (1, 2):
                                    cell.alignment = align_center
                                    if c_idx == 1 and isinstance(val, (datetime.datetime, pd.Timestamp)):
                                        cell.number_format = 'DD.MM.YYYY HH:mm:ss'
                                elif c_idx == 4:
                                    cell.alignment = align_right
                                    cell.number_format = '#,##0.00 "грн."'
                                else:
                                    cell.alignment = align_left
                            current_row += 1

                            # --- CHILD ROWS ---
                            total_cleared = 0.0
                            for _, match_row in group.iterrows():
                                dep_date = match_row['Дата поповнення']
                                dep_id = str(match_row['ID поповнення'])
                                dep_desc = str(match_row.get('Опис поповнення', '') or '')
                                dep_card = str(match_row.get('Картка отримувача', '') or dep_desc).strip()
                                cleared_amount = float(match_row.get('Сума компенсації', 0.0) or 0.0)
                                full_dep_amount = float(match_row.get('Сума поповнення', 0.0) or 0.0)
                                total_cleared += cleared_amount

                                dep_dt = pd.to_datetime(dep_date) if not isinstance(dep_date, (pd.Timestamp, datetime.datetime)) else dep_date
                                dep_date_str = dep_dt.strftime('%d.%m') if pd.notnull(dep_dt) else ""

                                if abs(cleared_amount - full_dep_amount) < 1e-5:
                                    detail_str = f"Повна компенсація: +{cleared_amount:g} грн від {dep_date_str}"
                                else:
                                    detail_str = f"Часткова компенсація: +{cleared_amount:g} грн із транзакції на +{full_dep_amount:g} грн від {dep_date_str}"

                                child_vals = [
                                    dep_date,
                                    dep_id,
                                    f"   ↳ Скомпенсовано на {dep_card}",
                                    None,
                                    cleared_amount,
                                    None,
                                    detail_str
                                ]
                                sheet.row_dimensions[current_row].height = 20
                                for c_idx, val in enumerate(child_vals, 1):
                                    cell = sheet.cell(row=current_row, column=c_idx, value=val)
                                    cell.font = child_font
                                    cell.fill = child_fill
                                    cell.border = header_border
                                    if c_idx in (1, 2):
                                        cell.alignment = align_center
                                        if c_idx == 1 and isinstance(val, (datetime.datetime, pd.Timestamp)):
                                            cell.number_format = 'DD.MM.YYYY HH:mm:ss'
                                    elif c_idx == 5:
                                        cell.alignment = align_right
                                        cell.number_format = '#,##0.00 "грн."'
                                    else:
                                        cell.alignment = align_left
                                current_row += 1

                            # --- SUMMARY ROW ---
                            net_expense = round(w_amount + total_cleared, 2)
                            summary_vals = [
                                None,
                                None,
                                "[!] РЕАЛЬНІ ВИТРАТИ ГОТІВКИ (ЧИСТИЙ КЕШ)",
                                None,
                                None,
                                net_expense,
                                "Списано на витрати леджера (чиста готівка)"
                            ]
                            sheet.row_dimensions[current_row].height = 22
                            for c_idx, val in enumerate(summary_vals, 1):
                                cell = sheet.cell(row=current_row, column=c_idx, value=val)
                                cell.font = summary_font
                                cell.fill = summary_fill
                                cell.border = header_border
                                if c_idx == 6:
                                    cell.alignment = align_right
                                    cell.number_format = '#,##0.00 "грн."'
                                else:
                                    cell.alignment = align_left
                            current_row += 1

                            # --- BLANK SEPARATOR ROW ---
                            sheet.row_dimensions[current_row].height = 12
                            for c_idx in range(1, 8):
                                sheet.cell(row=current_row, column=c_idx, value=None)
                            current_row += 1

                        elif event['type'] == 'twins':
                            t_row = event['data']
                            t_date = t_row['Дата зняття']
                            t_id_src = str(t_row['ID зняття'])
                            t_desc_src = str(t_row.get('Опис зняття', '') or '')
                            t_card_src = str(t_row.get('Картка джерела', '') or t_desc_src).strip()
                            t_id_dst = str(t_row['ID поповнення'])
                            t_desc_dst = str(t_row.get('Опис поповнення', '') or '')
                            t_card_dst = str(t_row.get('Картка отримувача', '') or t_desc_dst).strip()
                            t_amount = float(t_row.get('Сума зняття', 0.0) or 0.0)
                            t_cleared = float(t_row.get('Сума компенсації', 0.0) or 0.0)

                            dep_dt = pd.to_datetime(t_row['Дата поповнення']) if not isinstance(t_row['Дата поповнення'], (pd.Timestamp, datetime.datetime)) else t_row['Дата поповнення']
                            dep_date_str = dep_dt.strftime('%d.%m') if pd.notnull(dep_dt) else ""

                            twins_detail_str = f"Внутрішній переказ Twins: елімінація зустрічних транзакцій від {dep_date_str}"

                            twins_vals = [
                                t_date,
                                t_id_src,
                                f"Переказ {t_card_src} ➔ {t_card_dst}",
                                t_amount,
                                t_cleared,
                                0.00,
                                twins_detail_str
                            ]
                            sheet.row_dimensions[current_row].height = 22
                            for c_idx, val in enumerate(twins_vals, 1):
                                cell = sheet.cell(row=current_row, column=c_idx, value=val)
                                cell.font = twins_font
                                cell.fill = twins_fill
                                cell.border = header_border
                                if c_idx in (1, 2):
                                    cell.alignment = align_center
                                    if c_idx == 1 and isinstance(val, (datetime.datetime, pd.Timestamp)):
                                        cell.number_format = 'DD.MM.YYYY HH:mm:ss'
                                elif c_idx in (4, 5, 6):
                                    cell.alignment = align_right
                                    cell.number_format = '#,##0.00 "грн."'
                                else:
                                    cell.alignment = align_left
                            current_row += 1

                            # BLANK SEPARATOR ROW
                            sheet.row_dimensions[current_row].height = 12
                            for c_idx in range(1, 8):
                                sheet.cell(row=current_row, column=c_idx, value=None)
                            current_row += 1

                        elif event['type'] == 'transit_vika':
                            group = event['data']
                            first_row = group.iloc[0]
                            dep_date = first_row['Дата поповнення']
                            dep_id = str(first_row['ID поповнення'])
                            dep_desc = str(first_row.get('Опис поповнення', '') or '')
                            dep_card = str(first_row.get('Картка отримувача', '') or 'Monobank Sid Чорна').strip()
                            dep_amount = float(first_row.get('Сума поповнення', 0.0) or 0.0)

                            dep_dt = pd.to_datetime(dep_date) if not isinstance(dep_date, (pd.Timestamp, datetime.datetime)) else dep_date
                            dep_date_str = dep_dt.strftime('%d.%m') if pd.notnull(dep_dt) else ""

                            # PARENT ROW (Прихід транзиту від Віки)
                            parent_desc = f"Транзитні кошти Віки ({dep_card})"
                            parent_vals = [
                                dep_date,
                                dep_id,
                                parent_desc,
                                None,
                                dep_amount,
                                None,
                                f"Оригінальний прихід транзиту від {dep_date_str}"
                            ]
                            sheet.row_dimensions[current_row].height = 24
                            for c_idx, val in enumerate(parent_vals, 1):
                                cell = sheet.cell(row=current_row, column=c_idx, value=val)
                                cell.font = parent_font
                                cell.fill = parent_fill
                                cell.border = header_border
                                if c_idx in (1, 2):
                                    cell.alignment = align_center
                                    if c_idx == 1 and isinstance(val, (datetime.datetime, pd.Timestamp)):
                                        cell.number_format = 'DD.MM.YYYY HH:mm:ss'
                                elif c_idx == 5:
                                    cell.alignment = align_right
                                    cell.number_format = '#,##0.00 "грн."'
                                else:
                                    cell.alignment = align_left
                            current_row += 1

                            # CHILD ROWS (Компенсовані витрати інвестицій)
                            total_cleared = 0.0
                            for _, match_row in group.iterrows():
                                exp_date = match_row['Дата зняття']
                                exp_id = str(match_row['ID зняття'])
                                exp_desc = str(match_row.get('Опис зняття', '') or '')
                                cleared_amount = float(match_row.get('Сума компенсації', 0.0) or 0.0)
                                total_cleared += cleared_amount

                                exp_dt = pd.to_datetime(exp_date) if not isinstance(exp_date, (pd.Timestamp, datetime.datetime)) else exp_date
                                exp_date_str = exp_dt.strftime('%d.%m') if pd.notnull(exp_dt) else ""

                                detail_str = f"Компенсовано витрату інвестицій: -{cleared_amount:g} грн від {exp_date_str}"
                                child_vals = [
                                    exp_date,
                                    f"{dep_id}_clear_{exp_id}",
                                    f"   ↳ Покрито: [Транзит Віка] {exp_desc}",
                                    None,
                                    None,
                                    cleared_amount,
                                    detail_str
                                ]
                                sheet.row_dimensions[current_row].height = 20
                                for c_idx, val in enumerate(child_vals, 1):
                                    cell = sheet.cell(row=current_row, column=c_idx, value=val)
                                    cell.font = child_font
                                    cell.fill = child_fill
                                    cell.border = header_border
                                    if c_idx in (1, 2):
                                        cell.alignment = align_center
                                        if c_idx == 1 and isinstance(val, (datetime.datetime, pd.Timestamp)):
                                            cell.number_format = 'DD.MM.YYYY HH:mm:ss'
                                    elif c_idx == 6:
                                        cell.alignment = align_right
                                        cell.number_format = '#,##0.00 "грн."'
                                    else:
                                        cell.alignment = align_left
                                current_row += 1

                            # SUMMARY ROW (Вільний залишок транзиту)
                            rem_transit = round(dep_amount - total_cleared, 2)
                            if rem_transit > 0:
                                summary_vals = [
                                    None,
                                    None,
                                    "[!] ТРАНЗИТ ВІКА (ВІЛЬНИЙ ЗАЛИШОК)",
                                    None,
                                    rem_transit,
                                    None,
                                    f"Вільний залишок транзиту Віки: +{rem_transit:g} грн"
                                ]
                                sheet.row_dimensions[current_row].height = 22
                                for c_idx, val in enumerate(summary_vals, 1):
                                    cell = sheet.cell(row=current_row, column=c_idx, value=val)
                                    cell.font = summary_font
                                    cell.fill = summary_fill
                                    cell.border = header_border
                                    if c_idx == 5:
                                        cell.alignment = align_right
                                        cell.number_format = '#,##0.00 "грн."'
                                    else:
                                        cell.alignment = align_left
                                current_row += 1

                            # BLANK SEPARATOR ROW
                            sheet.row_dimensions[current_row].height = 12
                            for c_idx in range(1, 8):
                                sheet.cell(row=current_row, column=c_idx, value=None)
                            current_row += 1

                    sheet.freeze_panes = "A2"
                    if sheet.max_row >= 1 and sheet.max_column >= 1:
                        sheet.auto_filter.ref = sheet.dimensions

                else:
                    # Стилізація Daily_Dashboard
                    sheet.row_dimensions[1].height = 25
                    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

                    font_header = Font(bold=True, size=14, color="000000")
                    align_center = Alignment(horizontal="center", vertical="center")
                    thick_side = Side(style='thick', color="000000")
                    thin_side = Side(style='thin', color="A6A6A6")

                    # Дані з новими кольорами та вертикальним тонуванням (Зебра)
                    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    no_fill = PatternFill(fill_type=None)
                    rahom_fill = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
                    
                    for row_idx in range(1, sheet.max_row + 1):
                        is_header = (row_idx == 1)
                        curr_val = sheet.cell(row=row_idx, column=2).value
                        is_rahom = isinstance(curr_val, str) and curr_val.startswith('РАЗОМ')
                        is_even = row_idx % 2 == 0
                        
                        # Групування рядків за місяцями
                        if not is_header:
                            if is_rahom:
                                sheet.row_dimensions[row_idx].outline_level = 0
                            else:
                                sheet.row_dimensions[row_idx].outline_level = 1
                        
                        for col_idx in range(1, 9):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            
                            # Визначаємо межі для поточної клітинки
                            l_b = thin_side
                            r_b = thin_side
                            t_b = thin_side
                            b_b = thin_side

                            # Зовнішній контур всієї таблиці (Outline)
                            if col_idx == 1:
                                l_b = thick_side
                            if col_idx == 8:
                                r_b = thick_side
                            if row_idx == 1:
                                t_b = thick_side
                            if row_idx == sheet.max_row:
                                b_b = thick_side

                            # Обведення всього рядка заголовків
                            if row_idx == 1:
                                b_b = thick_side

                            # Колонки 'Дата' (ліва та права межі)
                            if col_idx == 2:
                                l_b = thick_side
                                r_b = thick_side

                            # Кожна з трьох колонок з різницями
                            if col_idx in (6, 7, 8):
                                l_b = thick_side
                                r_b = thick_side

                            # Жирні межі для РАЗОМ
                            if is_rahom:
                                t_b = thick_side
                                b_b = thick_side
                            
                            cell.border = Border(left=l_b, right=r_b, top=t_b, bottom=b_b)
                            
                            if is_header:
                                cell.font = font_header
                                cell.alignment = align_center
                                cell.fill = header_fill
                            else:
                                # Налаштування кольору фону (zebra)
                                if is_rahom:
                                    cell.fill = rahom_fill
                                elif is_even:
                                    cell.fill = zebra_fill
                                else:
                                    cell.fill = no_fill
                                        
                                # Встановлення ієрархії шрифтів
                                if is_rahom:
                                    if col_idx == 6:
                                        cell.font = Font(bold=True, italic=True, size=10, color="000000")
                                    else:
                                        cell.font = Font(bold=True, size=11, color="000000")
                                else:
                                    cell.font = Font(size=11, color="000000")

                                if col_idx in (1, 2):
                                    if is_rahom:
                                        cell.alignment = Alignment(horizontal="left", vertical="center")
                                    else:
                                        cell.alignment = Alignment(horizontal="center", vertical="center")
                                else:
                                    cell.alignment = Alignment(horizontal="right", vertical="center")

                    sheet.sheet_properties.outlinePr.summaryBelow = True
                    sheet.freeze_panes = "A2"
                    sheet.auto_filter.ref = sheet.dimensions

                # --- Налаштування ширини та форматів клітинок (для всіх вкладок) ---
                # 1. Знаходимо стовпчик «Категорія» та вираховуємо max_category_width як ліміт для даного листа
                category_col = None
                for col in sheet.columns:
                    header_val_str = str(col[0].value or '').strip()
                    if header_val_str in ('Категорія', COL_CAT):
                        category_col = col
                        break

                def _get_val_str(val, header_val, col_idx):
                    if val is None or pd.isna(val):
                        return ""
                    if isinstance(val, (int, float)):
                        import math
                        if math.isnan(val):
                            return ""
                        is_fin = (
                            header_val in (
                                'Сума', COL_AMOUNT, 'Залишок', COL_BALANCE, 'План', 'Витрати', 
                                'Дохід', 'Різниця за день', 'Різниця за місяць', 'Різниця загалом',
                                'Оригінальна сума', 'Скомпенсовано', 'Залишок (Витрати)'
                            ) or (sheet_name == 'Reconciliation_Audit' and col_idx in (4, 5, 6))
                        )
                        try:
                            return f"{val:,.2f}" if is_fin else (f"{int(val)}" if val == int(val) else str(val))
                        except Exception:
                            return str(val)
                    elif isinstance(val, (datetime.datetime, pd.Timestamp)):
                        return val.strftime('%d.%m.%Y %H:%M:%S')
                    elif isinstance(val, datetime.date):
                        return val.strftime('%d.%m.%Y')
                    else:
                        return str(val)

                if category_col is not None:
                    cat_max_len = 0
                    cat_header_val = str(category_col[0].value or '').strip()
                    for cell in category_col:
                        val_s = _get_val_str(cell.value, cat_header_val, category_col[0].column)
                        if len(val_s) > cat_max_len:
                            cat_max_len = len(val_s)
                    max_category_width = max(cat_max_len + 4, 12)
                else:
                    max_category_width = 35

                for col in sheet.columns:
                    col_letter = get_column_letter(col[0].column)
                    header_cell = col[0]
                    header_val = header_cell.value
                    col_idx = col[0].column
                    
                    max_len = 0
                    for cell in col:
                        val_s = _get_val_str(cell.value, header_val, col_idx)
                        if len(val_s) > max_len:
                            max_len = len(val_s)
                    
                    dynamic_width = max_len + 4
                    col_width = min(dynamic_width, max_category_width)
                    col_width = max(col_width, 12)
                    sheet.column_dimensions[col_letter].width = col_width

                    # Налаштування форматів чисел
                    start_r = 2
                    for row_idx in range(start_r, sheet.max_row + 1):
                        cell = sheet.cell(row=row_idx, column=col[0].column)
                        if header_val in (COL_DATE, 'Дата', 'Дата зняття', 'Дата поповнення', 'Дата відправки/зняття', 'Дата отримання/поповнення'):
                            cell.number_format = 'DD.MM.YYYY HH:mm:ss' if header_val not in ('Дата') or sheet_name != 'Daily_Dashboard' else 'DD.MM.YYYY'
                        elif header_val in (COL_AMOUNT, 'Сума зняття', 'Сума поповнення', 'Сума компенсації', 'Залишок зняття', 'Сума джерела', 'Сума отримувача', 'Залишок джерела'):
                            cell.number_format = '#,##0.00' if header_val != COL_AMOUNT else '[Color 10]#,##0.00;-#,##0.00;0.00'
                        elif header_val == COL_BALANCE:
                            cell.number_format = '#,##0.00'
                        elif header_val in ('План', 'Витрати', 'Дохід'):
                            cell.number_format = '#,##0.00'
                        elif header_val in ('Різниця за день', 'Різниця за місяць', 'Різниця загалом'):
                            cell.number_format = '[Color 10]#,##0.00;[Red]-#,##0.00;0.00'

        logger.info(f"Базу оновлено успішно. Разом транзакцій: {len(df)}")
    except PermissionError:
        logger.error(f"Файл {OUTPUT_FILE} відкритий. Будь ласка, закрийте Excel.")
    except Exception as e:
        logger.error(f"Помилка при збереженні: {e}", exc_info=True)
