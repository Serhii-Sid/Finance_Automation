import re
import hashlib
import logging
import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional
from config import (
    OUTPUT_FOLDER, OUTPUT_FILE, COL_ID, COL_DATE, COL_CAT, COL_CARD, 
    COL_DESC, COL_AMOUNT, COL_BALANCE, COL_MCC, IBAN_TO_CARD_MAP, CARDS_DICTIONARY, USEFUL_COLUMNS,
    MCC_MAP, CATEGORIES_KEYWORDS, INCOME_CATEGORIES, AMBIGUOUS_CATEGORIES
)

REGEX_CAT_CLEAN = r'^([^→•·*]{3,})(?=\s*[→•·*])' # Для Альянс Банку: витягує категорію з початку опису

logger = logging.getLogger(__name__)

ALLOWED_CATEGORIES = {
    # Витрати
    'Витрати на авто', 'Громадський транспорт', 'Житло та Побут', 'Зв\'язок', 
    'Здоров\'я та краса', 'Одяг та взуття', 'Зняття готівки', 'Компослуги', 
    'Переказ на чужий рахунок', 'Переказ на власний рахунок', 'Поштові послуги', 
    'Продукти', 'Ресторани та Розваги', 'Спорт', 'Техніка і цифрові товари', 
    'Оплата Кредитів', 'Інші витрати',
    # Надходження
    'Бонуси та Кешбек', 'Зарплата і виплати', 'Інші надходження', 
    'Переказ з власного рахунку', 'Переказ з чужого рахунку', 'Поповнення готівкою'
}

def detect_account_identity(text: str) -> Optional[str]:
    """
    Універсальний пошук IBAN або 4-значного хвоста карти в тексті з жорсткою ієрархією:
    1. Пріоритет №1 (IBAN): Пошук регулярним виразом UA + 27 цифр, перевірка в IBAN_TO_CARD_MAP.
    2. Пріоритет №2 (4 цифри): Пошук будь-якої послідовності з 4-х цифр поспіль, перевірка в CARDS_DICTIONARY.
    """
    if not text:
        return None

    # Очищуємо пробіли та переводимо в верхній регістр
    clean_text = re.sub(r'\s+', '', text).upper()

    # 1. Пріоритет №1 (IBAN)
    iban_match = re.search(r'UA\d{27}', clean_text)
    if iban_match:
        iban_code = iban_match.group(0)
        if iban_code in IBAN_TO_CARD_MAP:
            logger.info(f"Детектор (IBAN): знайдено IBAN {iban_code} -> '{IBAN_TO_CARD_MAP[iban_code]}'")
            return IBAN_TO_CARD_MAP[iban_code]
        # Вилучаємо знайдений IBAN з тексту пошуку, щоб його цифри не заважали резервному кроку
        clean_text = clean_text.replace(iban_code, '')

    # 2. Пріоритет №2 (Цифровий хвіст - Резерв)
    # Шукаємо всі 4-значні послідовності цифр в тексті
    digits4 = re.findall(r'\d{4}', clean_text)
    for d in digits4:
        if d in CARDS_DICTIONARY:
            logger.info(f"Детектор (Резерв): знайдено 4 цифри {d} -> '{CARDS_DICTIONARY[d]}'")
            return CARDS_DICTIONARY[d]

    return None

def standardize_df(df: pd.DataFrame) -> pd.DataFrame:
    """Приведення DataFrame до єдиного стандарту колонок."""
    if df is None or df.empty: return df
    
    # 1. Очищення та приведення до нижнього регістру для пошуку
    df.columns = [" ".join(str(c).split()).lower() for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]

    # 2. Мапінг варіацій назв у канонічні константи з config.py
    canonical_map = {
        COL_ID.lower(): COL_ID,
        "дата": COL_DATE, "дата і час": COL_DATE, "дата і час операції": COL_DATE, "дата операції": COL_DATE,
        "сума": COL_AMOUNT, "сума у валюті картки": COL_AMOUNT, "сума (uah)": COL_AMOUNT, "сума у валюті карти (uah)": COL_AMOUNT,
        "залишок": COL_BALANCE, "баланс": COL_BALANCE, "залишок після операції": COL_BALANCE, "залишок на кінець періоду": COL_BALANCE,
        "опис": COL_DESC, "опис операції": COL_DESC, "деталі": COL_DESC, "деталі операції": COL_DESC,
        "картка": COL_CARD, "номер картки": COL_CARD,
        "mcc": COL_MCC, "мсс": COL_MCC, "категорія": COL_CAT
    }
    
    # Перейменовуємо лише ті, що знайшли
    df = df.rename(columns={k: v for k, v in canonical_map.items() if k in df.columns})

    # 3. Логіка визначення категорії
    if COL_CAT not in df.columns:
        df[COL_CAT] = pd.NA
    else:
        # Очищаємо колонку від строкових плейсхолдерів порожніх значень
        df[COL_CAT] = df[COL_CAT].astype(str).str.strip().replace(
            ['nan', 'None', 'nan nan', '', '-', '<NA>'], pd.NA
        )

    # Заповнення пропусків у картці
    if COL_CARD in df.columns:
        placeholders = ['nan', 'None', '-', '', 'nan nan', 'x', 'X', 'х', 'Х', 'Невідома картка']
        df[COL_CARD] = df[COL_CARD].astype(str).str.strip().replace(placeholders, pd.NA)
        df[COL_CARD] = df[COL_CARD].ffill().bfill().fillna('Невідома картка')
    else:
        df[COL_CARD] = 'Невідома картка'

    if COL_BALANCE not in df.columns:
        df[COL_BALANCE] = 0.0

    # Гарантуємо наявність усіх цільових колонок, щоб уникнути KeyError
    for col in USEFUL_COLUMNS:
        if col not in df.columns:
            df[col] = pd.NA
    return df

def _normalize_text(s: pd.Series) -> pd.Series:
    """
    Крок 0: Нормалізація опису.
    Видаляє спецсимволи, цифри, слова-шум та приводить до нижнього регістру.
    Зберігає маркер 'plus380' для ідентифікації поповнень мобільних телефонів.
    """
    # 1. Нижній регістр та видалення спецсимволів
    clean = s.astype(str).str.lower()
    # 1.1. Зберігаємо +380 як тимчасовий маркер БЕЗ ЦИФР ('plusua'), щоб він вижив після strip digits
    clean = clean.str.replace(r'\+\s*380', 'plusua', regex=True)
    clean = clean.str.replace(r'[+•·→●○.,*/\-\\]+', ' ', regex=True)
    # 2. Видалення валют та технічних слів (шуму)
    clean = clean.str.replace(r'\b(uah|грн|оплата|покупка|переказ|списання)\b', ' ', regex=True)
    # 3. Видалення цифр (коди авторизації тощо) — 'plusua' не містить цифр, тому зберігається
    clean = clean.str.replace(r'\d+', ' ', regex=True)
    # 3.1. Відновлюємо фінальний маркер 'plus380' для пошуку за KEYWORD_MAP
    clean = clean.str.replace('plusua', 'plus380', regex=False)
    # 4. Видалення зайвих пробілів
    return clean.str.replace(r'\s+', ' ', regex=True).str.strip()

def classify_transfer(desc: str, amount: float, own_card: str) -> str:
    """
    Класифікує переказ на власний чи чужий рахунок.
    """
    # 1. Знаходимо всі 4-значні послідовності цифр в описі
    desc_digits = re.findall(r'(?<!\d)\d{4}(?!\d)', desc)
    
    # 2. Визначаємо суфікси власної картки
    own_suffixes = set()
    for d in re.findall(r'\d{4}', str(own_card)):
        own_suffixes.add(d)
    for d, name in CARDS_DICTIONARY.items():
        if name == own_card or name in str(own_card) or str(own_card) in name:
            own_suffixes.add(d)
            
    # 3. Виключаємо суфікси власної карти
    other_digits = [d for d in desc_digits if d not in own_suffixes]
    
    # 4. Перевіряємо, чи є серед інших цифр суфікс якоїсь іншої нашої карти
    has_other_own_card = False
    for d in other_digits:
        if d in CARDS_DICTIONARY:
            has_other_own_card = True
            break
            
    if has_other_own_card:
        return 'Переказ на власний рахунок' if amount < 0 else 'Переказ з власного рахунку'
    else:
        return 'Переказ на чужий рахунок' if amount < 0 else 'Переказ з чужого рахунку'

def clean_and_transform(df: pd.DataFrame) -> pd.DataFrame:
    """Фінальна векторизована обробка даних за принципом 'Keyword First'."""
    for col in [COL_AMOUNT, COL_BALANCE]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(r'\s+', '', regex=True).str.replace(',', '.', regex=False)
            df[col] = df[col].str.extract(r'(-?\d+\.?\d*)')[0]
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Попередня корекція знаку для технічних категорій (якщо банк не вказав мінус)
    expense_keywords = ['Покупка', 'Списання', 'Комісія', 'Видача готівки', 'Оплата']
    if COL_CAT in df.columns and COL_AMOUNT in df.columns:
        # Тимчасово приводимо до строки для маски
        mask = df[COL_CAT].astype(str).str.contains('|'.join(expense_keywords), case=False, na=False)
        df.loc[mask & (df[COL_AMOUNT] > 0), COL_AMOUNT] *= -1

    # --- Лійка рішень (Decision Funnel) за принципом "Keyword First" ---
    if COL_CAT in df.columns:
        # – Нормалізуємо початкову категорію банку (якщо вона є)
        df[COL_CAT] = df[COL_CAT].astype(str).str.strip().replace(
            ['nan', 'None', 'nan nan', '', '-', '<NA>'], pd.NA
        )

        # Створюємо тимчасову серію для класифікації
        final_cat = pd.Series(pd.NA, index=df.index, dtype='object')

        # 1. Нормалізація опису (викликається перед пошуком за ключовими словами)
        desc_norm = _normalize_text(df[COL_DESC]) if COL_DESC in df.columns else pd.Series()

        # ======================================================================
        # ПРІОРИТЕТ №1 (Найвищий): Пошук за ключовими словами (CATEGORIES_KEYWORDS)
        # ======================================================================
        if not desc_norm.empty:
            for cat, keywords in CATEGORIES_KEYWORDS.items():
                for kw in keywords:
                    pattern = rf'\b{re.escape(kw.lower())}'
                    mask = desc_norm.str.contains(pattern, na=False, regex=True)
                    if cat in ['Переказ на власний рахунок', 'Переказ з власного рахунку']:
                        # Враховуємо знак суми для внутрішніх переказів
                        for idx in df[mask & final_cat.isna()].index:
                            try:
                                amount = float(df.loc[idx, COL_AMOUNT] or 0)
                            except (ValueError, TypeError):
                                amount = 0.0
                            final_cat.loc[idx] = 'Переказ на власний рахунок' if amount < 0 else 'Переказ з власного рахунку'
                    else:
                        final_cat.loc[mask & final_cat.isna()] = cat

        # --- Супутні правила на основі опису ---
        # Крок 1.1: Розумні перекази (Власний vs Чужий)
        if COL_DESC in df.columns and COL_CARD in df.columns and COL_AMOUNT in df.columns:
            transfer_keywords = ['переказ', 'to card', 'на картку']
            
            # P2P патерни: тільки цифри/зірочки або тільки ім'я власника
            is_only_card = df[COL_DESC].astype(str).str.match(r'^[\d\s*]+$', na=False)
            name_regex = r'^[A-ZА-ЯІЄЇ][a-zа-яієїi\']+(\s+[A-ZА-ЯІЄЇ]\.?|\s+[A-ZА-ЯІЄЇ][a-zа-яієїi\']+)?$'
            is_only_name = df[COL_DESC].astype(str).str.match(name_regex, na=False)
            name_init_pattern = r'(?:[A-ZА-ЯІЄЇ][a-zа-яієїi\']+\s+[A-ZА-ЯІЄЇ]\.?|[A-ZА-ЯІЄЇ]\.?\s+[A-ZА-ЯІЄЇ][a-zа-яієїi\']+)'
            has_name_initial = df[COL_DESC].astype(str).str.contains(name_init_pattern, na=False, regex=True)
            
            for idx, row in df.iterrows():
                desc = str(row.get(COL_DESC, ''))
                desc_lower = desc.lower()
                
                is_transfer = any(tk in desc_lower for tk in transfer_keywords) or \
                              is_only_card.loc[idx] or \
                              is_only_name.loc[idx] or \
                              has_name_initial.loc[idx]
                              
                # Або якщо категорія вже визначена як переказ/P2P
                current_cat = final_cat.loc[idx]
                if not pd.isna(current_cat) and ('переказ' in str(current_cat).lower() or 'p2p' in str(current_cat).lower()):
                    is_transfer = True
                    
                if is_transfer and pd.isna(final_cat.loc[idx]):
                    try:
                        amount = float(row.get(COL_AMOUNT, 0) or 0)
                    except (ValueError, TypeError):
                        amount = 0.0
                        
                    own_card = str(row.get(COL_CARD, ''))
                    final_cat.loc[idx] = classify_transfer(desc, amount, own_card)

        # ======================================================================
        # ПРІОРИТЕТ №2 (Резервний): Категорія від банку або MCC
        # ======================================================================
        # 2.1. Категорія від MCC
        mcc_cats = pd.Series(pd.NA, index=df.index, dtype='object')
        if COL_MCC in df.columns:
            mcc_series = df[COL_MCC].astype(str).str.extract(r'(\d{4})')[0]
            mcc_cats = mcc_series.map(MCC_MAP)

        # 2.2. Категорія з опису за допомогою REGEX_CAT_CLEAN (для Альянс банку)
        desc_regex_cats = pd.Series(pd.NA, index=df.index, dtype='object')
        if COL_DESC in df.columns:
            desc_regex_cats = df[COL_DESC].astype(str).str.extract(REGEX_CAT_CLEAN)[0].str.strip()

        # Об'єднуємо резервні категорії:
        fallback_cat = df[COL_CAT].fillna(mcc_cats).fillna(desc_regex_cats)
        final_cat = final_cat.fillna(fallback_cat)

        # Записуємо фінальний результат назад у DataFrame
        df[COL_CAT] = final_cat

        # Визначаємо набір фінальних категорій для Bank Remapping
        final_categories = set(MCC_MAP.values()) | set(CATEGORIES_KEYWORDS.keys()) | \
                          {'Зв\'язок', 'Переказ на власний рахунок', 'Переказ з власного рахунку', 
                           'Переказ на чужий рахунок', 'Переказ з чужого рахунку'}

        # --- Додатковий крок: Bank Remapping (Очищення/перейменування категорій банку) ---
        for cat, keywords in CATEGORIES_KEYWORDS.items():
            for kw in keywords:
                is_not_standardized = ~df[COL_CAT].isin(final_categories) | df[COL_CAT].isin(AMBIGUOUS_CATEGORIES)
                mask = is_not_standardized & \
                       df[COL_CAT].astype(str).str.contains(kw, case=False, na=False, regex=False) & \
                       (~df[COL_CAT].astype(str).str.contains('власний рахунок', na=False))
                df.loc[mask, COL_CAT] = cat

        # --- Контекстні правила (Врахування знаку суми) ---
        if COL_AMOUNT in df.columns:
            # Розділення власних переказів (якщо потрапили через CATEGORIES_KEYWORDS або старі категорії)
            mask_internal = df[COL_CAT].isin([
                'Переказ між власними рахунками', 
                'Переказ на власний рахунок', 
                'Переказ з власного рахунку',
                'Переказ на свою картку'
            ])
            df.loc[mask_internal & (df[COL_AMOUNT] < 0), COL_CAT] = 'Переказ на власний рахунок'
            df.loc[mask_internal & (df[COL_AMOUNT] > 0), COL_CAT] = 'Переказ з власного рахунку'

        # --- Фінальний Fallback (Перевірка білого списку) ---
        whitelist_mask = ~df[COL_CAT].isin(ALLOWED_CATEGORIES) | df[COL_CAT].isna()
        if COL_AMOUNT in df.columns:
            df.loc[whitelist_mask & (df[COL_AMOUNT] < 0), COL_CAT] = 'Інші витрати'
            df.loc[whitelist_mask & (df[COL_AMOUNT] >= 0), COL_CAT] = 'Інші надходження'
        else:
            df.loc[whitelist_mask, COL_CAT] = 'Інші витрати'

    if COL_DATE in df.columns:
        if not pd.api.types.is_datetime64_any_dtype(df[COL_DATE]):
            df[COL_DATE] = df[COL_DATE].astype(str).str.replace('\xa0', ' ', regex=False)
            df[COL_DATE] = df[COL_DATE].str.replace(r'\s+', ' ', regex=True).str.strip()
    
    df[COL_DATE] = pd.to_datetime(df[COL_DATE], dayfirst=True, errors='coerce')
    df[COL_CARD] = df[COL_CARD].astype(str).replace(CARDS_DICTIONARY)
    
    return df[df[COL_DATE].notna()].copy()

def make_short_id_vectorized(df: pd.DataFrame) -> pd.Series:
    """Генерація унікальних ID для транзакцій."""
    combined = (
        df[COL_DATE].dt.strftime('%Y-%m-%d %H:%M:%S') + "_" + 
        df[COL_CARD].astype(str) + "_" + 
        df[COL_AMOUNT].map('{:.2f}'.format)
    )
    return combined.apply(lambda x: str(int(hashlib.md5(x.encode()).hexdigest()[:8], 16) % 10**8).zfill(8))

def rotate_outputs():
    """Керування версіями файлу Ledger."""
    base_name = "Total_Ledger_v{}.xlsx"
    for i in [2, 1]:
        old_v = os.path.join(OUTPUT_FOLDER, base_name.format(i))
        new_v = os.path.join(OUTPUT_FOLDER, base_name.format(i+1))
        if os.path.exists(old_v): os.replace(old_v, new_v)
    if os.path.exists(OUTPUT_FILE): 
        os.replace(OUTPUT_FILE, os.path.join(OUTPUT_FOLDER, base_name.format(1)))

def reconcile_and_merge(new_data: pd.DataFrame, df_base: pd.DataFrame) -> pd.DataFrame:
    """Порівнює нові дані з існуючою базою, уникаючи дублікатів."""
    if df_base.empty:
        new_data[COL_ID] = make_short_id_vectorized(new_data)
        return new_data

    if COL_ID in df_base.columns:
        df_base[COL_ID] = df_base[COL_ID].astype(str).str.zfill(8)

    if not pd.api.types.is_datetime64_any_dtype(df_base[COL_DATE]):
        df_base[COL_DATE] = pd.to_datetime(df_base[COL_DATE], dayfirst=True, errors='coerce')
        
    df_base[COL_AMOUNT] = pd.to_numeric(df_base[COL_AMOUNT], errors='coerce')
    df_base[COL_BALANCE] = pd.to_numeric(df_base[COL_BALANCE], errors='coerce')
    
    keys = [COL_DATE, COL_CARD, COL_DESC, COL_AMOUNT]
    df_base_keys = df_base[keys].drop_duplicates()
    
    merged = new_data.merge(df_base_keys, on=keys, how='left', indicator=True)
    unique_new = new_data[merged['_merge'] == 'left_only'].copy()
    
    count_new = len(unique_new)
    count_dupes = len(new_data) - count_new
    
    if not unique_new.empty:
        logger.info(f"Звірка: додано {count_new} нових транзакцій. Відфільтровано {count_dupes} дублікатів.")
        unique_new[COL_ID] = make_short_id_vectorized(unique_new)
        return pd.concat([df_base, unique_new], ignore_index=True)
    
    return df_base

def save_final_ledger(df: pd.DataFrame, script_path: str):
    """Зберігає фінальний файл Excel та створює бекап скрипта."""
    try:
        rotate_outputs()
        # Створюємо копію для експорту та додаємо допоміжну колонку періоду
        df_export = df.copy()
        # Додаємо порядковий номер на початок
        df_export.insert(0, '№ п/п', range(1, len(df_export) + 1))
        
        ukr_months = {
            1: 'Січень', 2: 'Лютий', 3: 'Березень', 4: 'Квітень', 5: 'Травень', 6: 'Червень',
            7: 'Липень', 8: 'Серпень', 9: 'Вересень', 10: 'Жовтень', 11: 'Листопад', 12: 'Грудень'
        }
        df_export.insert(2, 'Місяць', df_export[COL_DATE].dt.month.map(ukr_months) + " " + df_export[COL_DATE].dt.year.astype(str))

        # Підготовка даних для роздільних таблиць
        df_income = df_export[df_export[COL_AMOUNT] > 0].copy()
        df_expenses = df_export[df_export[COL_AMOUNT] < 0].copy()

        # Видаляємо непотрібні колонки для листа "Аналіз"
        cols_to_drop_analysis = ['№ п/п', 'Місяць']
        df_income = df_income.drop(columns=cols_to_drop_analysis, errors='ignore')
        df_expenses = df_expenses.drop(columns=cols_to_drop_analysis, errors='ignore')

        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', datetime_format='dd.mm.yyyy hh:mm:ss') as writer:
            # Записуємо основний реєстр
            df_export.to_excel(writer, index=False, sheet_name='Реєстр')
            
            # Записуємо роздільні таблиці на окремі аркуші
            df_expenses.to_excel(writer, sheet_name='Витрати', index=False)
            df_income.to_excel(writer, sheet_name='Надходження', index=False)

            # Стилізація обох листів
            for sheet_name in ['Реєстр', 'Витрати', 'Надходження']:
                sheet = writer.sheets[sheet_name]
                sheet.row_dimensions[1].height = 30  # Збільшуємо висоту шапки

                # --- Стилізація заголовків ---
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="5A5A5A", end_color="5A5A5A", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                thin_side = Side(style='thin', color="000000")
                header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                # Знаходимо всі комірки заголовків
                for row in sheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        if cell.value is not None:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = header_border

                # --- Стилізація даних (Сітка та Зебра) ---
                thin_grid_side = Side(style='thin', color="A6A6A6")
                thick_border_side = Side(style='medium', color="000000")
                # data_border = Border(left=thin_grid_side, right=thin_grid_side, top=thin_grid_side, bottom=thin_grid_side)
                zebra_fill = PatternFill(start_color="E9E9E9", end_color="E9E9E9", fill_type="solid")
                data_alignment = Alignment(vertical="center")

                prev_period = None
                group_start = 2
                    
                # Проходимо по всіх рядках даних (починаючи з 2-го)
                for row_idx in range(2, sheet.max_row + 1):
                    curr_period = None
                    is_new_period = False

                    if sheet_name == 'Реєстр':
                        # Колонка "Місяць" для "Реєстру" знаходиться за індексом 3 (№ п/п, ID_Транзакції, Місяць)
                        curr_period = sheet.cell(row=row_idx, column=3).value
                        is_new_period = prev_period is not None and curr_period != prev_period

                    is_even = row_idx % 2 == 0

                    # Визначаємо, чи змінився місяць для малювання межі та групування
                    is_new_period = prev_period is not None and curr_period != prev_period
                    
                    if is_new_period and sheet_name == 'Реєстр':
                        # Групуємо попередній місяць: всі рядки КРІМ першого (group_start)
                        if row_idx - 1 > group_start:
                            sheet.row_dimensions.group(group_start + 1, row_idx - 1, outline_level=1)
                        group_start = row_idx

                    for col_idx in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)

                        # Якщо новий період у Реєстрі - додаємо жирну лінію зверху
                        if sheet_name == 'Реєстр':
                            top_s = thick_border_side if is_new_period else thin_grid_side
                        else: # Для листа "Аналіз" завжди використовуємо тонку сітку
                            top_s = thin_grid_side
                        cell.border = Border(left=thin_grid_side, right=thin_grid_side, top=top_s, bottom=thin_grid_side)
                        cell.alignment = data_alignment
                        if is_even:
                            cell.fill = zebra_fill
                    
                    prev_period = curr_period

                # Групуємо останній період: всі рядки КРІМ першого
                if sheet_name == 'Реєстр' and sheet.max_row > group_start:
                    sheet.row_dimensions.group(group_start + 1, sheet.max_row, outline_level=1)

                sheet.sheet_properties.outlinePr.summaryBelow = False # Кнопки "+" зверху над групами
                sheet.freeze_panes = "A2"       # Закріплюємо верхній рядок
                sheet.auto_filter.ref = sheet.dimensions # Додаємо автофільтри

                # --- Налаштування форматів та ширини (цикл по колонках) ---
                fixed_widths = {
                    COL_ID: 10, COL_DATE: 20, # '№ п/п' та 'Місяць' видалені з df_income/df_expenses
                    COL_CAT: 18, COL_CARD: 30, COL_DESC: 15, COL_AMOUNT: 12, COL_BALANCE: 12
                }

                for col_idx in range(1, sheet.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    header_val = sheet.cell(row=1, column=col_idx).value
                    
                    if header_val in fixed_widths:
                        sheet.column_dimensions[col_letter].width = fixed_widths[header_val]

                    # Форматування чисел та дат у колонках
                    for row_idx in range(2, sheet.max_row + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if header_val == COL_DATE:
                            cell.number_format = 'DD.MM.YYYY HH:mm:ss'
                        elif header_val == COL_AMOUNT:
                            cell.number_format = '[Color 10]#,##0.00;-#,##0.00;0.00'
                        elif header_val == COL_BALANCE:
                            cell.number_format = '#,##0.00'

        logger.info(f"Базу оновлено успішно. Разом транзакцій: {len(df)}")
    except PermissionError:
        logger.error(f"Файл {OUTPUT_FILE} відкритий. Будь ласка, закрийте Excel.")
    except Exception as e:
        logger.error(f"Помилка при збереженні: {e}", exc_info=True)

def apply_bank_specific_post_processing(df: pd.DataFrame) -> pd.DataFrame:
    """
    Застосовує коригування для специфічних банків (наприклад, ПУМБ).
    Тут ми рахуємо накопичувальний баланс, якщо він не надається банком.
    """
    pumb_cards = [v for v in CARDS_DICTIONARY.values() if 'ПУМБ' in v]
    if not pumb_cards:
        return df

    df = df.copy()
    df[COL_BALANCE] = df[COL_BALANCE].astype(float)
    
    for card_name in pumb_cards:
        if card_name in df[COL_CARD].values:
            mask = df[COL_CARD] == card_name
            # Сортуємо від старих до нових для розрахунку cumsum
            df_card = df[mask].sort_values(by=COL_DATE)
            df.loc[df_card.index, COL_BALANCE] = df_card[COL_AMOUNT].fillna(0).cumsum()
            
    return df

def detect_internal_transfers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Інтелектуальне розпізнавання внутрішніх переказів між власними рахунками.
    Аналізує весь підсумковий DataFrame перед збереженням.

    Етап 1 (За масками): Шукає в описі будь-які 4 цифри, що збігаються з останніми 
    цифрами карток (0550, 7854, 4169, 0707, 0627, 4258, 4882, 6333, 4805, 0950).
    Якщо знайдено — присвоює категорію 'Переказ на власний рахунок' (якщо сума від'ємна)
    або 'Переказ з власного рахунку' (якщо сума додатна).

    Етап 2 (За близнюками): Шукає пари транзакцій з однаковою сумою (різними знаками)
    та однаковим часом (з точністю до хвилини). Маркує їх як внутрішні перекази.
    """
    if df is None or df.empty:
        return df

    df = df.copy()

    # 0. Пріоритет брендів: Класифікуємо/захищаємо категорію брендів перед перевірками
    # Це гарантує, що ключові слова з CATEGORIES_KEYWORDS мають найвищий пріоритет
    desc_norm = _normalize_text(df[COL_DESC]) if COL_DESC in df.columns else pd.Series()
    if not desc_norm.empty:
        final_cat = df[COL_CAT].copy()
        protected_cats = set(CATEGORIES_KEYWORDS.keys()) - {'Переказ на власний рахунок', 'Переказ з власного рахунку'}
        
        for cat, keywords in CATEGORIES_KEYWORDS.items():
            for kw in keywords:
                pattern = rf'\b{re.escape(kw.lower())}'
                mask = desc_norm.str.contains(pattern, na=False, regex=True)
                
                # Дозволяємо перезапис тільки якщо поточне значення не є захищеною категорією
                can_overwrite = ~final_cat.isin(protected_cats) | final_cat.isna()
                
                if cat in ['Переказ на власний рахунок', 'Переказ з власного рахунку']:
                    # Враховуємо знак суми для внутрішніх переказів
                    for idx in df[mask & can_overwrite].index:
                        try:
                            amount = float(df.loc[idx, COL_AMOUNT] or 0)
                        except (ValueError, TypeError):
                            amount = 0.0
                        final_cat.loc[idx] = 'Переказ на власний рахунок' if amount < 0 else 'Переказ з власного рахунку'
                else:
                    final_cat.loc[mask & can_overwrite] = cat
        df[COL_CAT] = final_cat

    # Створюємо набір кандидатів для аналізу (Зв'язок та інші бренд-категорії не чіпаємо)
    candidate_cats = set(AMBIGUOUS_CATEGORIES) | {
        'Перекази', 'Переказ коштів', 'Переказ між власними рахунками',
        'Переказ на власний рахунок', 'Переказ з власного рахунку',
        'Переказ на чужий рахунок', 'Переказ з чужого рахунку',
        'Інші витрати', 'Інші надходження'
    }

    # Маски (останні 4 цифри власних карток)
    target_suffixes = {'0550', '7854', '4169', '0707', '0627', '4258', '4882', '6333', '4805', '0950'}

    changed_mask = 0
    changed_twin = 0
    validated_own_indices = set()

    # Етап 1 (За масками)
    # Скануємо всі рядки, які є кандидатами
    if COL_CAT in df.columns:
        candidates_idx = df[df[COL_CAT].isin(candidate_cats)].index
    else:
        candidates_idx = df.index

    for idx in candidates_idx:
        row = df.loc[idx]
        desc = str(row.get(COL_DESC, ''))
        own_card = str(row.get(COL_CARD, ''))
        
        # Визначаємо суфікси власної картки
        own_suffixes = set()
        for d in re.findall(r'\d{4}', own_card):
            own_suffixes.add(d)
        for d, name in CARDS_DICTIONARY.items():
            if name == own_card or name in own_card or own_card in name:
                own_suffixes.add(d)

        # Шукаємо standalone 4-значні числа в описі
        # Використовуємо regex, щоб уникнути витягування 4-значних частин з довших чисел (як IBAN)
        digits4 = re.findall(r'(?<!\d)\d{4}(?!\d)', desc)
        matched_suffix = None
        for d in digits4:
            if d in target_suffixes and d not in own_suffixes:
                matched_suffix = d
                break
        
        if matched_suffix:
            try:
                amount = float(row.get(COL_AMOUNT, 0) or 0)
            except (ValueError, TypeError):
                amount = 0.0
                
            new_cat = 'Переказ на власний рахунок' if amount < 0 else 'Переказ з власного рахунку'
            
            validated_own_indices.add(idx)
            
            if df.loc[idx, COL_CAT] != new_cat:
                df.loc[idx, COL_CAT] = new_cat
                changed_mask += 1
                logger.info(
                    f"detect_internal_transfers (Етап 1: Маска): idx={idx}, "
                    f"desc='{desc[:60]}', amount={amount:.2f}, matched_suffix='{matched_suffix}' → '{new_cat}'"
                )

    # Етап 2 (За близнюками)
    # Шукаємо пари транзакцій з однаковою абсолютною сумою (з різними знаками) та різницею часу <= 5 хвилин
    # Не об'єднуємо транзакції, якщо вони відбулися по одній картці.
    
    valid_mask = df[COL_AMOUNT].notna() & df[COL_DATE].notna()
    valid_df = df[valid_mask].copy()
    valid_df['abs_amount'] = valid_df[COL_AMOUNT].abs().round(2)
    valid_df = valid_df[valid_df['abs_amount'] > 0]
    
    overwrite_cats = set(AMBIGUOUS_CATEGORIES) | {
        'Перекази', 'Переказ коштів', 'Переказ між власними рахунками',
        'Переказ на власний рахунок', 'Переказ з власного рахунку',
        'Переказ на чужий рахунок', 'Переказ з чужого рахунку',
        'Інші витрати', 'Інші надходження'
    }
    
    twin_indices = set()
    grouped = valid_df.groupby('abs_amount')
    for abs_amt, group in grouped:
        if len(group) < 2:
            continue
        
        pos_trans = group[group[COL_AMOUNT] > 0].copy()
        neg_trans = group[group[COL_AMOUNT] < 0].copy()
        
        if pos_trans.empty or neg_trans.empty:
            continue
            
        def to_naive(t):
            return t.replace(tzinfo=None) if t.tzinfo is not None else t
            
        pos_trans['naive_date'] = pos_trans[COL_DATE].apply(to_naive)
        neg_trans['naive_date'] = neg_trans[COL_DATE].apply(to_naive)
        
        pos_trans = pos_trans.sort_values('naive_date')
        neg_trans = neg_trans.sort_values('naive_date')
        
        matched_pos_indices = set()
        
        for neg_idx, neg_row in neg_trans.iterrows():
            best_pos_idx = None
            best_time_diff = pd.Timedelta(days=99999)
            
            neg_time = neg_row['naive_date']
            neg_card = neg_row[COL_CARD]
            
            for pos_idx, pos_row in pos_trans.iterrows():
                if pos_idx in matched_pos_indices:
                    continue
                if pos_row[COL_CARD] == neg_card:
                    continue
                    
                pos_time = pos_row['naive_date']
                time_diff = abs(neg_time - pos_time)
                
                if time_diff <= pd.Timedelta(minutes=5):
                    if time_diff < best_time_diff:
                        best_time_diff = time_diff
                        best_pos_idx = pos_idx
                        
            if best_pos_idx is not None:
                matched_pos_indices.add(best_pos_idx)
                twin_indices.add((neg_idx, best_pos_idx))

    for neg_idx, pos_idx in twin_indices:
        validated_own_indices.add(neg_idx)
        validated_own_indices.add(pos_idx)
        
        # Для витрати (neg)
        row_neg = df.loc[neg_idx]
        current_cat_neg = row_neg.get(COL_CAT)
        if pd.isna(current_cat_neg) or current_cat_neg in overwrite_cats:
            if df.loc[neg_idx, COL_CAT] != 'Переказ на власний рахунок':
                df.loc[neg_idx, COL_CAT] = 'Переказ на власний рахунок'
                changed_twin += 1
                logger.info(
                    f"detect_internal_transfers (Етап 2: Близнюки витрата): idx={neg_idx}, "
                    f"desc='{str(row_neg.get(COL_DESC, ''))[:60]}', amount={row_neg.get(COL_AMOUNT):.2f} → 'Переказ на власний рахунок'"
                )
                
        # Для доходу (pos)
        row_pos = df.loc[pos_idx]
        current_cat_pos = row_pos.get(COL_CAT)
        if pd.isna(current_cat_pos) or current_cat_pos in overwrite_cats:
            if df.loc[pos_idx, COL_CAT] != 'Переказ з власного рахунку':
                df.loc[pos_idx, COL_CAT] = 'Переказ з власного рахунку'
                changed_twin += 1
                logger.info(
                    f"detect_internal_transfers (Етап 2: Близнюки дохід): idx={pos_idx}, "
                    f"desc='{str(row_pos.get(COL_DESC, ''))[:60]}', amount={row_pos.get(COL_AMOUNT):.2f} → 'Переказ з власного рахунку'"
                )

    # Етап 3 (Демоція непідтверджених власних переказів)
    own_transfer_cats = {
        'Переказ на власний рахунок', 
        'Переказ з власного рахунку',
        'Переказ між власними рахунками',
        'Переказ на свою картку'
    }
    
    demoted_count = 0
    for idx, row in df.iterrows():
        current_cat = row.get(COL_CAT)
        if current_cat in own_transfer_cats:
            if idx not in validated_own_indices:
                try:
                    amount = float(row.get(COL_AMOUNT, 0) or 0)
                except (ValueError, TypeError):
                    amount = 0.0
                
                new_cat = 'Переказ на чужий рахунок' if amount < 0 else 'Переказ з чужого рахунку'
                df.loc[idx, COL_CAT] = new_cat
                demoted_count += 1
                logger.info(
                    f"detect_internal_transfers (Демоція): idx={idx}, "
                    f"desc='{str(row.get(COL_DESC, ''))[:60]}', amount={amount:.2f} → '{new_cat}'"
                )

    logger.info(
        f"detect_internal_transfers завершено: змінено за масками {changed_mask}, "
        f"змінено за близнюками {changed_twin}, демонтовано {demoted_count} транзакцій."
    )

    # Фінальний Fallback (Перевірка білого списку) для гарантії відповідності ALLOWED_CATEGORIES
    whitelist_mask = ~df[COL_CAT].isin(ALLOWED_CATEGORIES) | df[COL_CAT].isna()
    if COL_AMOUNT in df.columns:
        df.loc[whitelist_mask & (df[COL_AMOUNT] < 0), COL_CAT] = 'Інші витрати'
        df.loc[whitelist_mask & (df[COL_AMOUNT] >= 0), COL_CAT] = 'Інші надходження'
    else:
        df.loc[whitelist_mask, COL_CAT] = 'Інші витрати'

    return df

def adjust_balance_with_credit_limit(df: pd.DataFrame, credit_limit: float) -> pd.DataFrame:
    """
    Adjusts the balance column by subtracting the credit limit.
    Assumes COL_BALANCE is present and numeric-like.
    """
    if df is None or df.empty or COL_BALANCE not in df.columns or credit_limit <= 0:
        return df

    # Make a copy to avoid SettingWithCopyWarning
    df_copy = df.copy()

    def adjust(val):
        try:
            s = str(val).replace('\xa0', '').replace(' ', '').replace(',', '.')
            m = re.search(r'(-?\d+\.?\d*)', s)
            if m:
                return float(m.group(1)) - credit_limit
            return val
        except Exception:
            return val
    df_copy[COL_BALANCE] = df_copy[COL_BALANCE].apply(adjust)
    return df_copy